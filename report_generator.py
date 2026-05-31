"""
📊 Sales Report Generator
==========================
Generates Excel and PDF sales reports for admin.
"""

import os
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def generate_excel_report(events_data, stats) -> bytes:
    """
    Generate Excel sales report.
    Returns bytes of the .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # ── Colors ──────────────────────────────────────────────
    DARK    = "0D0D0D"
    GOLD    = "C9A84C"
    GREEN   = "2ECC71"
    WHITE   = "F5F5F5"
    GRAY    = "2A2A2A"

    # ── Title ────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "TicketFlow — Sales Report"
    ws["A1"].font      = Font(name="Arial", bold=True, size=16, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font      = Font(name="Arial", size=10, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Summary Stats ────────────────────────────────────────
    ws.merge_cells("A4:G4")
    ws["A4"] = "SUMMARY"
    ws["A4"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A4"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A4"].alignment = Alignment(horizontal="center")

    summary = [
        ("Total Confirmed Bookings", stats["total_confirmed"]),
        ("Total Pending",            stats["total_pending"]),
        ("Total Cancelled",          stats["total_cancelled"]),
        ("Total Revenue",            f"EGP {stats['total_revenue']:,.0f}"),
    ]
    for i, (label, value) in enumerate(summary, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = str(value)
        ws[f"A{i}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"B{i}"].font = Font(name="Arial", size=10, color="2ECC71")

    # ── Events Table ─────────────────────────────────────────
    headers = ["Event", "Date", "Venue", "Capacity", "Sold", "Available", "Revenue (EGP)"]
    row = 10
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "EVENT BREAKDOWN"
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=GRAY)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor="C9A84C")
        cell.alignment = Alignment(horizontal="center")

    row += 1
    for i, e in enumerate(events_data):
        available = e["capacity"] - e["booked"]
        data_row = [
            e["name"], e["date"], e["venue"],
            e["capacity"], e["confirmed_count"],
            available, f"{e['total_revenue']:,.0f}"
        ]
        fill_color = "1A1A1A" if i % 2 == 0 else "222222"
        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = Font(name="Arial", size=10, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
        row += 1

    # ── Column widths ────────────────────────────────────────
    widths = [30, 20, 20, 12, 10, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_report(events_data, stats) -> str:
    """Generate PDF sales report. Returns file path."""
    os.makedirs("temp_reports", exist_ok=True)
    filepath = f"temp_reports/report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    doc    = SimpleDocTemplate(filepath, pagesize=A4,
                                topMargin=15*mm, bottomMargin=15*mm,
                                leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle("title", fontSize=20, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#C9A84C"), alignment=TA_CENTER)
    story.append(Paragraph("TicketFlow — Sales Report", title_style))
    story.append(Spacer(1, 5*mm))

    sub_style = ParagraphStyle("sub", fontSize=9, textColor=colors.HexColor("#888888"),
                                 alignment=TA_CENTER)
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))
    story.append(Spacer(1, 8*mm))

    # Summary table
    summary_data = [
        ["Metric", "Value"],
        ["✅ Confirmed Bookings", str(stats["total_confirmed"])],
        ["⏳ Pending",            str(stats["total_pending"])],
        ["❌ Cancelled",          str(stats["total_cancelled"])],
        ["💰 Total Revenue",      f"EGP {stats['total_revenue']:,.0f}"],
    ]
    st = Table(summary_data, colWidths=[90*mm, 60*mm])
    st.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#C9A84C")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.black),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND",  (0, 1), (-1, -1), colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR",   (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#1A1A1A"), colors.HexColor("#222222")]),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#2A2A2A")),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ]))
    story.append(st)
    story.append(Spacer(1, 8*mm))

    # Events breakdown
    h_style = ParagraphStyle("h", fontSize=12, fontName="Helvetica-Bold",
                               textColor=colors.HexColor("#C9A84C"))
    story.append(Paragraph("Event Breakdown", h_style))
    story.append(Spacer(1, 4*mm))

    ev_headers = ["Event", "Date", "Sold", "Capacity", "Revenue"]
    ev_data    = [ev_headers]
    for e in events_data:
        ev_data.append([
            e["name"][:25], e["date"][:15],
            str(e["confirmed_count"]), str(e["capacity"]),
            f"EGP {e['total_revenue']:,.0f}"
        ])

    et = Table(ev_data, colWidths=[60*mm, 35*mm, 20*mm, 20*mm, 35*mm])
    et.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#C9A84C")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.black),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1),
         [colors.HexColor("#1A1A1A"), colors.HexColor("#222222")]),
        ("TEXTCOLOR",     (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#2A2A2A")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(et)

    doc.build(story)
    return filepath
